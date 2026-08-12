"""바나나몰 엑셀 파서 — DESIGN.md 2.1.

68열이지만 열 **순서를 믿지 않는다.** 헤더 이름으로 찾는다. 순서로 읽으면
다운로드 양식이 한 번만 바뀌어도 조용히 엉뚱한 값을 집는다.

찾지 못한 헤더는 조용히 넘어가지 않고 이름을 들고 올라온다. 1000개를 돌리기 전에
"이 파일에는 이런 헤더가 있다"를 사람이 볼 수 있어야 한다.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field

#: 우리가 쓰는 이름 → 엑셀에 나올 법한 헤더 후보들.
FIELDS: dict[str, tuple[str, ...]] = {
    "code": ("상품번호", "상품코드", "goodsno", "품번"),
    "name": ("상품명", "제품명", "goodsnm"),
    "brand": ("브랜드", "제조사", "brand"),
    "maker": ("거래처", "공급사", "제조사명"),
    "price": ("판매가", "판매가격", "price"),
    "cost": ("원가", "정가", "소비자가"),
    "category": ("카테고리", "분류"),
    "thumb": ("대표이미지url", "대표이미지", "목록이미지url", "목록이미지"),
    "body": ("상세설명", "상세정보", "상품상세", "detail"),
    "url": ("상품url", "상품링크"),
}

OPTION_HEADERS = ("옵션1", "옵션2", "옵션3", "옵션")


def norm(s) -> str:
    """헤더 비교용 정규화 — 공백·괄호·기호를 지우고 소문자로."""
    return re.sub(r"[\s()\[\]/_\-.]", "", str(s or "")).lower()


@dataclass
class Option:
    name: str
    values: list[str] = field(default_factory=list)


@dataclass
class Row:
    code: str = ""
    name: str = ""
    brand: str = ""
    maker: str = ""
    price: str = ""
    cost: str = ""
    category: str = ""
    thumb: str = ""
    body: str = ""
    url: str = ""
    options: list[Option] = field(default_factory=list)
    #: 엑셀에서 읽은 원본 행 전체. 못 쓴 열을 나중에 확인할 수 있게 남긴다.
    raw: dict = field(default_factory=dict)

    @property
    def option_values(self) -> list[str]:
        return [v for o in self.options for v in o.values]


def parse_option(cell: str) -> Option | None:
    """`옵션명=값,값,값` 형식을 푼다 (2.1).

    7장 — 원본 접두 번호(`01. `)는 벗긴다. 그대로 두면 출력 쪽 번호와 겹친다.
    """
    text = str(cell or "").strip()
    if not text:
        return None
    name, sep, rest = text.partition("=")
    if not sep:
        name, rest = "", text
    values = []
    for v in rest.split(","):
        v = re.sub(r"^\s*\d+\s*[.)]\s*", "", v.strip())
        if v:
            values.append(v)
    if not values:
        return None
    return Option(name=name.strip(), values=values)


def _header_map(header: list) -> dict[str, int]:
    """헤더 행 → {우리 이름: 열 번호}."""
    seen = {norm(h): i for i, h in enumerate(header) if str(h or "").strip()}
    out: dict[str, int] = {}
    for key, candidates in FIELDS.items():
        for cand in candidates:
            i = seen.get(norm(cand))
            if i is not None:
                out[key] = i
                break
    return out


def _find_header_row(rows: list[list], limit: int = 10) -> int:
    """헤더가 첫 줄이 아닐 수 있다. 아는 이름이 가장 많이 걸리는 줄을 고른다."""
    best, best_hits = 0, -1
    for i, row in enumerate(rows[:limit]):
        hits = len(_header_map(row))
        if hits > best_hits:
            best, best_hits = i, hits
    return best


@dataclass
class Sheet:
    rows: list[Row]
    headers: list[str]
    #: 찾지 못한 항목. UI 에 그대로 보여준다.
    missing: list[str]


def load(path_or_bytes) -> Sheet:
    """엑셀(.xlsx)을 읽어 행 목록으로."""
    import io

    import openpyxl

    src = io.BytesIO(path_or_bytes) if isinstance(path_or_bytes, (bytes, bytearray)) else path_or_bytes
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    if not grid:
        raise ValueError("빈 시트다")

    hi = _find_header_row(grid)
    header = [str(h or "").strip() for h in grid[hi]]
    cols = _header_map(grid[hi])
    missing = [k for k in ("code", "name", "body") if k not in cols]

    opt_cols = [
        i for i, h in enumerate(grid[hi])
        if any(norm(h) == norm(o) for o in OPTION_HEADERS)
    ]

    rows: list[Row] = []
    for raw in grid[hi + 1 :]:
        if not any(str(c or "").strip() for c in raw):
            continue
        get = lambda k: str(raw[cols[k]] or "").strip() if k in cols and cols[k] < len(raw) else ""  # noqa: E731
        row = Row(
            code=get("code"), name=get("name"), brand=get("brand"), maker=get("maker"),
            price=get("price"), cost=get("cost"), category=get("category"),
            thumb=get("thumb"), body=get("body"), url=get("url"),
            raw={header[i]: raw[i] for i in range(min(len(header), len(raw))) if header[i]},
        )
        for i in opt_cols:
            if i < len(raw):
                opt = parse_option(raw[i])
                if opt:
                    row.options.append(opt)
        if row.code or row.name:
            rows.append(row)

    wb.close()
    return Sheet(rows=rows, headers=[h for h in header if h], missing=missing)
